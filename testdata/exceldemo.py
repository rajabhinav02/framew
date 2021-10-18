import openpyxl

book = openpyxl.load_workbook("C:\\Test Data\\TestExcel.xlsx")
sheet= book.active
Dict = {}
d1=sheet.cell(row=1, column=1)
#p#rint(d1.value)
#sheet.cell(row=4, column=5).value = "Mo"
#print(sheet.cell(row=4, column=5).value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value== "TC2":
        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print([Dict])
#print(sheet['E4'].value)
