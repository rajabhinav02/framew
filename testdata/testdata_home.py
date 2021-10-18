import openpyxl

class formtestdata:

    formdata = [{"firstname": "Abhi", "email": "ghi", "pwd":"raj" }, {"firstname": "ms", "email": "uip", "pwd":"rpj"}]

    @staticmethod
    def testdata(testname):
        Dict = {}
        book=openpyxl.load_workbook("C:\\Test Data\\TestExcel.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value== testname:
                for j in range(2, sheet.max_column+1):
                    Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i,column=j).value

        return [Dict]

